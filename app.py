import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def read_file(file_name):
    return xl.load_workbook(file_name)


def get_sheet(workbook, sheet_name):
    return workbook[sheet_name]


def get_existing_amount_of_rows(sheet):
    return sheet.max_row


def insert_new_row_with_value(sheet, row, column_index, corrected_value):
    sheet.cell(row, column_index).value = corrected_value


def get_new_corrected_value(sheet, row, column_index):
    cell = sheet.cell(row, column_index)
    return cell.value * 0.9


def create_columns_with_values(sheet, row_start_index, column_index_original_value, column_index_new_value):
    for row in range(row_start_index, get_existing_amount_of_rows(sheet) + 1):
        corrected_value = get_new_corrected_value(sheet, row, column_index_original_value)
        insert_new_row_with_value(sheet, row, column_index_new_value, corrected_value)


def get_values_from_given_range(sheet, min_row, min_column, max_column):
    return Reference(sheet,
                     min_row=min_row,
                     max_row=get_existing_amount_of_rows(sheet),
                     min_col=min_column,
                     max_col=max_column)


def create_chart(sheet, start_row_index, get_column_index_new_place):
    chart = BarChart()
    values = get_values_from_given_range(sheet, start_row_index, get_column_index_new_place, get_column_index_new_place)
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')


fileName = input('What should be the file name? ')
wb = read_file(f'assets/transactions.xlsx')
sheet = get_sheet(wb, 'Sheet1')
row_start_index = 2
column_index_original_value = 3
column_index_new_value = column_index_original_value + 1

create_columns_with_values(sheet, row_start_index, column_index_original_value, column_index_new_value)
create_chart(sheet, row_start_index, column_index_new_value)

wb.save(f'assets/new_{fileName}.xlsx')
print(f'Successfully created the file: new_{fileName}.xlsx')
